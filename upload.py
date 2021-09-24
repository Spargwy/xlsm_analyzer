#!/usr/bin/env python
from openpyxl import load_workbook
import sqlite3
import sys


def read_xlsm_file(file_path):
    wb_obj = load_workbook(file_path)
    sheet_obj = wb_obj.active

    max_rows = sheet_obj.max_row
    max_columns = sheet_obj.max_column
    non_void_columns = []
    for column in range(1, max_columns + 1):
        column_name = sheet_obj.cell(row=1, column=column)
        if column_name.value is None:
            continue
        non_void_columns.append(column)

    non_void_rows = []
    for row in range(2, max_rows + 1):
        endpoint_id = sheet_obj.cell(row=row, column=1)
        if endpoint_id.value is None:
            continue
        non_void_rows.append(row)

    table_values = {}
    for row in non_void_rows:
        for column in non_void_columns:
            endpoint_name = sheet_obj.cell(row=row, column=column)
            endpoint_id = sheet_obj.cell(row=row, column=1)
            if endpoint_name.value is None:
                endpoint_name.value = 'void'
            table_values[row] = {endpoint_id.value: endpoint_name.value}
    insert_data_from_file(table_values)


def insert_data_from_file(table_values):
    db_name = 'excel_files_db'
    create_tables(db_name)
    con = sqlite3.connect(db_name)
    cur = con.cursor()
    sql = '''INSERT INTO excel_table(id, endpoint_id, endpoint_name) VALUES(?, ?, ?)'''
    for table_sheet in table_values:
        for endpoint_id in table_values[table_sheet]:
            data = (table_sheet, endpoint_id, table_values[table_sheet][endpoint_id])
            cur.execute(sql, data)
    con.commit()


def create_tables(db_name):
    con = sqlite3.connect(db_name)
    cur = con.cursor()

    cur.execute('''CREATE TABLE IF NOT EXISTS excel_table(id integer, endpoint_id integer, endpoint_name text)''')
    con.commit()
    con.close()


def main():
    args = sys.argv
    try:
        file_path = args[1]
        read_xlsm_file(file_path)
    except Exception as e:
        print("error: ", e)
        print("Enter correct arguments")


main()
